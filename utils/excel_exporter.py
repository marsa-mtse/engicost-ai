import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

def export_boq_branded(df: pd.DataFrame, company_name: str, project_name: str = "مشروع تسعير", author_name: str = "المهندس") -> io.BytesIO:
    """
    Exports a pandas DataFrame to a beautifully formatted Excel file using openpyxl.
    Adds company headers, sums, and professional styling.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.sheet_view.rightToLeft = True # RTL Support for Arabic

    # --- Header Styling ---
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Blue
    header_font = Font(name="Cairo", size=16, bold=True, color="FFFFFF")
    sub_font = Font(name="Cairo", size=12, bold=True, color="1E293B")
    
    # Border
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 1. Add Company Branding Row
    ws.merge_cells('A1:F2')
    ws['A1'] = f"🏗️ {company_name or 'EngiCost AI'}"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # 2. Add Project Info Info
    ws['A4'] = "اسم المشروع:"
    ws['B4'] = project_name
    ws['A4'].font = sub_font
    
    ws['D4'] = "التاريخ:"
    ws['E4'] = datetime.datetime.now().strftime("%Y-%m-%d")
    ws['D4'].font = sub_font

    ws['A5'] = "إعداد بواسطة:"
    ws['B5'] = author_name
    ws['A5'].font = sub_font

    # 3. Add DataFrame Data
    start_row = 7
    # Write column headers
    cols = list(df.columns)
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = Font(name="Cairo", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid") # EngiCost Blue
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name="Cairo")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if type(value) != str else "right")
            
            # Format numbers
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    last_row = start_row + len(df)

    # 4. Add Automatically calculated SUM row if there is a 'Total' or 'الإجمالي' column
    # Let's find numeric columns to sum
    sum_row = last_row + 1
    ws.cell(row=sum_row, column=1, value="الإجمالي الكلي").font = Font(name="Cairo", bold=True)
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=2)
    
    for c_idx, col_name in enumerate(cols, 1):
        if 'سعر' in str(col_name) or 'إجمالي' in str(col_name) or 'total' in str(col_name).lower():
            # Add excel formula =SUM(...)
            col_letter = ws.cell(row=1, column=c_idx).column_letter
            formula = f"=SUM({col_letter}{start_row + 1}:{col_letter}{last_row})"
            cell = ws.cell(row=sum_row, column=c_idx, value=formula)
            cell.font = Font(name="Cairo", bold=True)
            cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            cell.number_format = '#,##0.00'
            cell.border = thin_border

    # 5. Add Signature Block at the bottom
    sig_row = sum_row + 3
    ws.cell(row=sig_row, column=2, value="توقيع المهندس المقيم:").font = sub_font
    ws.cell(row=sig_row, column=4, value="توقيع الاستشاري:").font = sub_font
    
    # 6. Adjust Column Widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except: pass
        adjusted_width = (max_length + 5)
        ws.column_dimensions[column].width = min(adjusted_width, 40) # Max 40 width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
